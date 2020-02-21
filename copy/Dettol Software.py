
from tkinter import *
import sqlite3
import tkinter.messagebox

import datetime
import math
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os
from openpyxl import Workbook
from openpyxl.worksheet.pagebreak import Break
import random
from openpyxl import Workbook
conn = sqlite3.connect("vi.db")
c = conn.cursor()

date = datetime.datetime.now().date()
today = date.today()
d2 = str(today.strftime("%d-%m-%Y"))
products_list = []
product_price = []
product_quantity = []
product_id = []
product_cgst=[]
product_cgstp=[]
product_sgst=[]
product_sgstp=[]
product_mrp=[]
product_amount=[]
product_dis=[]
product_rate=[]
product_code=[]
labels_list = []
akm=[]
prolist=[]
prolis=[]
custlis=[]
proli=[]
custlist=[]
p4=""
p1=""
p2=""
p3=""
sck="nil"
yi=""
li=""
il=""
f=[]
g=[]
h=[]
y=[]
class Application:
    def __init__(self, master, *args, **kwargs):
        self.master = master
        self.left = Frame(master, width=650, height=768, bg='teal')
        self.left.pack(side=LEFT)
        self.right = Frame(master, width=716, height=768, bg='lightblue')
        self.right.pack(side=RIGHT)


        
        tkvar = StringVar(self.left)
        scrollbar = Scrollbar(self.left)
        con=sqlite3.Connection('vi.db')
        cur=con.cursor()
        query = "SELECT pdescription FROM prod "
        cur.execute(query)
        res = cur.fetchall()
        self.valuel=""
    
        for row in res:
            prolist.append(row)
        for row in prolist:
            prolis.append(row[0])

        def update_list(self):
            try:
                search_term = self.left.search_var.get()
                lbox_list =prolis
                self.left.lbox.delete(0, END)
                for item in lbox_list:
                    if search_term.lower() in item.lower():
                        self.left.lbox.insert(END, item)
            except:
                print("")
        def CurSelet(evt):
            try:
                self.valuel=str((self.left.lbox.get(self.left.lbox.curselection())))
                self.entername=self.valuel
            except:
                print("")
            #print (self.valuel)
        def selectlist1(event):
            try:
                listbox = event.widget
                if listbox['height'] <3:
                    self.left.lbox.config(height=5,bg='lightblue',width=24,fg='black')
                else:
                    self.left.lbox.config(height=1,bg='teal',width=3,fg='teal')
                if listbox['height'] <3:
                    self.left.lbox2.config(height=5,bg='lightblue',width=24,fg='black')
                else:
                    self.left.lbox2.config(height=1,bg='teal',width=3,fg='teal')
            except:
                print("b")
        self.left.search_var = StringVar()
        self.left.search_var.trace("w", lambda name, index, mode: update_list(self))
        self.left.entry = Entry(self.left, textvariable=self.left.search_var,width=13, font=('arial 18 bold'), bg='lightblue')
        self.left.entry.place(x=190, y=80)
        self.left.entry.focus()
        #self.left.entry = Entry(self.left, textvariable=self.left.search_var, width=25)
        self.left.lbox = Listbox(self.left, width=3, height=1,yscrollcommand = scrollbar.set,bg='teal',fg='teal')
        self.left.lbox.place(x=370,y=84)
        self.left.lbox.bind('<<ListboxSelect>>',CurSelet)
        self.left.bind("<Double-Button-1>", selectlist1)
        Scrollbar(self.left.lbox, orient="vertical")
        scrollbar.config( command = self.left.lbox.yview )
        scrollbar.pack( fill="y")
        scrollbar.place(x=520,y=115)
        update_list(self)
        def selectk(event):
            listbox = event.widget
            self.left.lbox.config(height=6,width=24,bg='lightblue',fg='black')
        self.left.entry.bind("<Key>",selectk)
        scrollbar2 = Scrollbar(self.left)
        con=sqlite3.Connection('vi.db')
        cur=con.cursor()
        query = "SELECT name FROM cust "
        cur.execute(query)
        ree = cur.fetchall()
        for row in ree:
            custlist.append(row)
        for row in custlist:
            proli.append(row[0])
        def update_lists(self):
            try:
                search_term = self.left.searchi_var.get()
                lbox2_list =proli
                self.left.lbox2.delete(0, END)
                for item in lbox2_list:
                    if search_term.lower() in item.lower():
                        self.left.lbox2.insert(END, item)
            except:
                print("b")
        self.left.searchi_var = StringVar()
        self.left.searchi_var.trace("w", lambda name, index, mode: update_lists(self))
        self.entercust = Entry(self.left, textvariable=self.left.searchi_var, width=30, font=('arial 10 bold'), bg='lightblue')
        self.entercust.place(x=225, y=245)
        self.entercust.focus()
        def CurSelets(evt):
            try:

                self.valuels=str((self.left.lbox2.get(self.left.lbox2.curselection())))
                self.entercust.delete(0, END)
                self.entercust.insert(12,self.valuels)
            except:
                print("b")
        self.left.lbox2 = Listbox(self.left, width=3, height=1,yscrollcommand = scrollbar2.set,bg='teal',fg='teal')
        self.left.lbox2.place(x=441,y=253)
        self.left.lbox2.bind('<<ListboxSelect>>',CurSelets)
        self.left.bind("<Double-Button-1>", selectlist1)

        Scrollbar(self.left.lbox2, orient="vertical")
        scrollbar2.config( command = self.left.lbox2.yview )
        scrollbar2.pack( fill="y")
        scrollbar2.place(x=585,y=245)
        update_lists(self)
        def selectks(event):
            try:
                listbox = event.widget
                self.left.lbox2.config(height=6,width=23,bg='lightblue',fg='black')
            except:
                print("b")
        self.entercust.bind("<Key>",selectks)

        self.heading = Label(self.left, text="        Shrinath  Agency          ", font=('arial 40 bold'), bg='red')
        self.heading.place(x=0, y=0)
        self.date_l = Label(self.right, text="Today's Date: " + str(d2), font=('arial 16 bold'), bg='lightblue', fg='red')
        self.date_l.place(x=0, y=0)
        self.tproduct = Label(self.right, text="Product", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.tproduct.place(x=0, y=60)
        self.tamount = Label(self.right, text="MRP", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.tamount.place(x=142, y=60)
        self.tamount = Label(self.right, text="DIS%", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.tamount.place(x=187, y=60)
        self.tax_amt = Label(self.right, text="Tax Amt", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.tax_amt.place(x=238, y=60)
        self.tquantity = Label(self.right, text="Quan", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.tquantity.place(x=298, y=60)
        self.gst = Label(self.right, text="  Rate", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.gst.place(x=342, y=60)
        self.gst = Label(self.right, text="CGST %", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.gst.place(x=402, y=60)
        self.gst = Label(self.right, text="CGST", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.gst.place(x=462, y=60)
        self.gst = Label(self.right, text="SGST %", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.gst.place(x=507, y=60)
        self.gst = Label(self.right, text="SGST", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.gst.place(x=567, y=60)
        self.gst = Label(self.right, text="Amount", font=('arial 10 bold'), bg='lightblue', fg='black')
        self.gst.place(x=620, y=60)
        self.change_l = Label(self.right, text="Given Amount", font=('arial 18 bold'), bg='red')
        self.change_l.place(x=0, y=650)
        self.change_e = Entry(self.right, width=25, font=('arial 18 bold'), bg='white')
        self.change_e.place(x=190, y=650)
        self.change_btn = Button(self.right,text="Calculate Change", width=18, height=2, bg='orange', command=self.change_func)
        self.change_btn.place(x=540, y=644)
        '''
        self.enterid = Label(self.left, text="Enter Hsncode", font=('arial 18 bold'), bg='teal')
        self.enterid.place(x=0, y=80)
        self.enteride = Entry(self.left, width=25, font=('arial 18 bold'), bg='lightblue')
        self.enteride.place(x=190, y=80)
        self.enteride.focus()
        '''
        self.enternam = Label(self.left, text="Enter Product Name", font=('arial 14 bold'), bg='teal')
        self.enternam.place(x=0, y=80)
        self.search_btn = Button(self.left, text="Search", width=10, height=4,font=('Algerian 13'), bg='orange', command=self.ajax)
        self.search_btn.place(x=550, y=80)
        #self.bill = Button(self.left, text="Databa", width=16, height=2,font=('arial 10'), bg='grey', fg='black', command=self.gen)
        #self.bill.place(x=540, y=590)
        self.bill = Button(self.left, text="Search \n customer", width=10, height=2,font=('arial 12'), bg='saddle brown', fg='slategray1', command=self.contact)
        self.bill.place(x=550, y=245)
        Button(self.left, text="Create User", compound='center', font=("arial", 12), bg='Blue', fg='white', width=10,height=2, command=lambda: self.gen()).place(x=550,y=295)
        self.entercusts= Label(self.left, text="Customer Name", font=('arial 18 bold'), bg='teal')
        self.entercusts.place(x=0, y=245)
        self.enteradd = Label(self.left, text="Customer Address", font=('arial 18 bold'), bg='teal')
        self.enteradd.place(x=0, y=285)
        self.enteradd = Entry(self.left, width=30, font=('arial 10 bold'), bg='lightblue')
        self.enteradd.place(x=225, y=288)
        self.GST =Label(self.left, text="Customer GST", font=('arial 18 bold'), bg='teal')
        self.GST.place(x=0, y=325)
        self.GST =Entry(self.left, width=30, font=('arial 10 bold'), bg='lightblue')
        self.GST.place(x=225, y=328)
        self.GST.focus()
        # fill it later by the function ajax
        self.productname = Label(self.left, text="", font=('arial 17 bold'), bg='teal', fg='black')
        self.productname.place(x=0, y=410)
        self.total_l = Label(self.right, text="", font=('arial 18 bold'), bg='lightblue', fg='black')
        self.total_l.place(x=0, y=560)
        #self.master.bind("<Return>", self.ajax)
        #self.master.bind("<Up>", self.add_to_cart)
        #self.master.bind("<space>", self.printers)
    def contact(self, *args, **kwargs):
        self.enteradd.delete(0, END)
        self.GST.delete(0, END)
        self.p1=self.entercust.get()
        self.p2=self.enteradd.get()
        self.p3=self.GST.get()

        query = "SELECT * FROM cust WHERE name=?"
        result = c.execute(query, (self.p1, ))
        for self.r in result:
            self.p1 = self.r[0]
            self.p2 = self.r[1]
            self.p4=self.r[2]
            self.p3 = self.r[3]
            f.clear()
            g.clear()
            h.clear()
            y.clear()
            f.append(self.p1)
            g.append(self.p2)
            h.append(self.p3)
            y.append(self.p4)
            #self.entercust.insert(2,str(self.r[0]))
            self.enteradd.insert(2,str(self.r[1]))
            self.GST.insert(2,str(self.r[3]))
    def gen(self, *args, **kwargs):
        root.destroy()
        root1=Tk()
        root1.title(" ADMIN ")
        root1.geometry('805x520')
        def account1():
            usr = user.get()
            pasw = pas.get()
            if(usr=="SHRINATH" and pasw=="SHRINATH"):
                def db():
                    con=sqlite3.Connection('vi.db')
                    cur=con.cursor()
                    #cur.execute("create table if not exists prod(HSNcode varchar(20) primary key,pdescription varchar(45),MRP number(8),rate number(8),cgst number(5),sgst number(8))")
                    l = (hc.get(),pd.get(),mr.get(),Gs.get())
                    cur.execute("insert into cust values(?,?,?,?)",l)
                    con.commit()
                    cur.execute("select * from cust")
                    kl=cur.fetchall()
                    tkinter.messagebox.showinfo("Congrats", "successfully accepted")
                root1.destroy()
                root4 = Tk()
                root4.title("Database")
                root4.geometry('930x645')
                Label(root4, text="Shrinath General Agency", font=("algerian", 22), fg="black", bg="red", width=52).place(x=0,y=40)
                Label(root4, text="Please fill all the details given below all are mandatory", font=("cooper black", 21),fg="red").place(x=0,y=100)
                Label(root4, text="Customer name", font=("cooper black", 15),fg=('Blue')).place(x=200,y=150)
                hc = Entry(width=20, font=("arial", 11), bg=('#ffcccc'))
                hc.place(x=400,y=150)
                Label(root4, text="customer address", font=("cooper black", 15),fg=('Blue')).place(x=200,y=190)
                pd = Entry(width=20, font=("arial", 11), bg=('#ffcccc'))
                pd.place(x=400,y=190)
                Label(root4, text="Mobile no.", font=("cooper black", 15),fg=('Blue')).place(x=200,y=230)
                mr = Entry(width=20, font=("arial", 11), bg=('#ffcccc'))
                mr.insert(END,"+91")
                mr.place(x=400,y=230)
                Label(root4, text="GST no.", font=("cooper black", 15),fg=('Blue')).place(x=200,y=270)
                Gs = Entry(width=20, font=("arial", 11), bg=('#ffcccc'))
                Gs.insert(END,"NOT PROVIDED")
                Gs.place(x=400,y=270)
                ''''
                Label(root4, text="CGST", font=("cooper black", 15),fg=('Blue')).place(x=200,y=310)
                cgst = Entry(width=20, font=("arial", 11), bg=('#ffcccc'))
                cgst.place(x=400,y=310)
                Label(root4, text="SGST", font=("cooper black", 15),fg=('Blue')).place(x=200,y=350)
                sgst = Entry(width=20, font=("arial", 11), bg=('#ffcccc'))
                sgst.place(x=400,y=350)
                '''
                Button(root4, text="submit", compound='center', font=("arial", 10), bg='Blue', fg='white', width=10, command=lambda: db()).place(x=430,y=420)
                root4.mainloop()
            else:
                showerror('Log In Failed', "Invalid Username or Password")
        Label(root1, text="Shrinath general store", font=("algerian", 22), fg="black", bg="red", width=52).grid(row=0, column=0, columnspan=4)
        Label(root1, text="Username: ", font=("cooper black", 15),fg=('Blue')).place(x=200,y=150)
        user = Entry(width=20,font=("arial", 11), bg=('#ffcccc'))
        user.insert(END,"SHRINATH")
        user.place(x=320,y=150)
        Label(root1, text="Password: ", font=("cooper black", 15),fg=('Blue')).place(x=200,y=190)
        pas = Entry(width=20,font=("arial", 11), bg=('#ffcccc'))
        pas.insert(END,"SHRINATH")
        pas.place(x=320,y=190)
        Button(root1, text="Log In", compound='center', font=("arial", 10), bg='Blue', fg='white', width=10, command=lambda: account1()).place(x=210,y=250)


    def ajax(self, *args, **kwargs):
        self.get_id = ""

        self.get_name = self.entername
        if self.get_id=="":
            query = "SELECT * FROM prod WHERE pdescription=?"
            result = c.execute(query, (self.get_name, ))
        #else:
            #query = "SELECT * FROM prod WHERE HSNcode=?"
            #result = c.execute(query, (self.get_id, ))
        for self.r in result:
            self.get_id = self.r[0]
            self.get_name = self.r[1]
            self.get_mrp = self.r[2]
            self.get_rate = self.r[3]
            self.cgst=self.r[4]
            self.sgst=self.r[5]
            self.productname.configure(text="Product = " + str(self.get_name)+"      MRP = "+str(self.get_mrp))
        #self.pprice.configure(text="Price=Rs." + str(self.get_rate))
        self.quantity_l = Label(self.left, text="Enter Quantity", font=('arial 18 bold'), bg='teal')
        self.quantity_l.place(x=0, y=475)
        self.quantity_e = Entry(self.left, width=15, font=('arial 18 bold'), bg='lightblue')
        self.quantity_e.place(x=190, y=475)
        self.quantity_e.focus()
        self.discount_l = Label(self.left, text="Enter Discount", font=('arial 18 bold'), bg='teal')
        self.discount_l.place(x=0, y=515)
        self.discount_e = Entry(self.left, width=15, font=('arial 18 bold'), bg='lightblue')
        self.discount_e.place(x=190, y=515)
        self.discount_e.insert(END, 0)
        # add to cart button
        self.add_to_cart_btn = Button(self.left, text="Add To Cart", width=18, height=4, bg='orange', command=self.add_to_cart)
        self.add_to_cart_btn.place(x=500, y=475)
        self.bill_btn = Button(self.left, text="Generate Bill", width=22, height=3,font=('arial 10 bold'), bg='red', fg='white', command=self.printers)
        self.bill_btn.place(x=180, y=600)

    def add_to_cart(self, *args, **kwargs):
        list2=[90,125,160,195,230,265,300,335,370,405,440,475,510,545,580,615]
        self.productname.configure(text="")

        self.quantity_value = int(self.quantity_e.get())
        if self.quantity_value >10000000:
            tkinter.messagebox.showinfo("Error", "Not that many products in our inventory.")
        else:
            x=float(self.quantity_value) * float(self.get_rate)
            self.final_price =x-(x*float(self.discount_e.get())/100)
            #self.final_price =x
            akm.append(x)
            #print(akm,x)
            self.camt = ((float(self.final_price) * float(self.cgst))/100)
            self.samt = ((float(self.final_price) * float(self.sgst))/100)
            self.amount=self.final_price+self.camt+self.samt
            product_code.append(self.get_id)
            products_list.append(self.get_name)
            product_price.append(float("%.2f"%self.final_price))
            product_quantity.append(self.quantity_value)
            product_id.append(self.get_id)
            product_cgst.append(self.cgst)
            product_mrp.append(float("%.2f"%self.get_mrp))
            product_cgstp.append(float("%.2f"%self.camt))
            product_sgst.append(self.sgst)
            product_sgstp.append(float("%.2f"%self.samt))
            product_amount.append(float("%.2f"%self.amount))
            product_dis.append(self.discount_e.get())
            product_rate.append(float("%.2f"%self.get_rate))
            self.x_index = 0
            self.y_index = 90
            self.counter = 0

            def account1(x,lp):
                try:
                    self.y_ind=lp
                    self.tempname = Label(self.right, text="                                             ", font=('arial 8 '), bg='lightblue', fg='black')
                    self.tempname.place(x=0, y=self.y_ind)
                    self.tempname.grid_forget()
                    self.tempdis = Label(self.right, text="                            ", font=('arial 9 bold'), bg='lightblue', fg='red')
                    self.tempdis.place(x=197, y=self.y_ind)
                    self.tempqt = Label(self.right, text="                           ", font=('arial 9 bold'), bg='lightblue', fg='black')
                    self.tempqt.place(x=298,y=self.y_ind)
                    self.temprate = Label(self.right, text="                          ", font=('arial 9 bold'), bg='lightblue', fg='black')
                    self.temprate.place(x=346, y=self.y_ind)
                    self.tempprice = Label(self.right, text="                         ",font=('arial 9 bold'), bg='lightblue', fg='black')
                    self.tempprice.place(x=238, y=self.y_ind)
                    self.tempcgst = Label(self.right, text="                          ", font=('arial 9 bold'), bg='lightblue', fg='green')
                    self.tempcgst.place(x=424, y=self.y_ind)
                    self.tempsgst = Label(self.right, text="                          ", font=('arial 9 bold'), bg='lightblue', fg='green')
                    self.tempsgst.place(x=525, y=self.y_ind)
                    self.tempmrp = Label(self.right, text="                           ", font=('arial 9 bold'), bg='lightblue', fg='black')
                    self.tempmrp.place(x=142, y=self.y_ind)
                    self.tempamount = Label(self.right, text="             ", font=('arial 9 bold'), bg='lightblue', fg='red')
                    self.tempamount.place(x=620, y=self.y_ind)
                    self.tempcgstp = Label(self.right, text="                         ", font=('arial 9 bold'), bg='lightblue', fg='black')
                    self.tempcgstp.place(x=462, y=self.y_ind)
                    self.tempsgstp = Label(self.right, text="              ", font=('arial 9 bold'), bg='lightblue', fg='black')
                    self.tempsgstp.place(x=567, y=self.y_ind)

                    for self.r in range(0,len(list2)):
                        if list2[self.r]==lp:
                            del products_list[self.r]
                            del product_dis[self.r]
                            del product_quantity[self.r]
                            del product_rate[self.r]
                            del product_price[self.r]
                            del product_cgst[self.r]
                            del product_sgst[self.r]
                            del product_mrp[self.r]
                            del product_amount[self.r]
                            del product_cgstp[self.r]
                            del product_sgstp[self.r]
                            del akm[self.r]
                except:
                    tkinter.messagebox.showerror('ERROR OCCURED ', "PLEASE RESTART THE SOFTWARE ELSE BILL BE WRONG")





            for self.p in products_list:
                self.tempname = Label(self.right, text=str(products_list[self.counter]), font=('arial 8 '), bg='lightblue', fg='black')
                self.tempname.place(x=0, y=self.y_index)
                labels_list.append(self.tempname)
                self.tempdis = Label(self.right, text=str(product_dis[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='red')
                self.tempdis.place(x=197, y=self.y_index)
                labels_list.append(self.tempdis)
                self.tempqt = Label(self.right, text=str(product_quantity[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='black')
                self.tempqt.place(x=298,y=self.y_index)
                labels_list.append(self.tempqt)
                self.temprate = Label(self.right, text=str(product_rate[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='black')
                self.temprate.place(x=346, y=self.y_index)
                labels_list.append(self.temprate)
                self.tempprice = Label(self.right, text=str(product_price[self.counter]),font=('arial 9 bold'), bg='lightblue', fg='black')
                self.tempprice.place(x=238, y=self.y_index)
                labels_list.append(self.tempprice)
                self.tempcgst = Label(self.right, text=str(product_cgst[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='green')
                self.tempcgst.place(x=424, y=self.y_index)
                labels_list.append(self.tempcgst)
                self.tempsgst = Label(self.right, text=str(product_sgst[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='green')
                self.tempsgst.place(x=525, y=self.y_index)
                labels_list.append(self.tempsgst)
                self.tempmrp = Label(self.right, text=str(product_mrp[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='black')
                self.tempmrp.place(x=142, y=self.y_index)
                labels_list.append(self.tempmrp)
                self.tempamount = Label(self.right, text=str(product_amount[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='red')
                self.tempamount.place(x=620, y=self.y_index)
                labels_list.append(self.tempamount)
                self.tempcgstp = Label(self.right, text=str(product_cgstp[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='black')
                self.tempcgstp.place(x=462, y=self.y_index)
                labels_list.append(self.tempcgstp)
                self.tempsgstp = Label(self.right, text=str(product_sgstp[self.counter]), font=('arial 9 bold'), bg='lightblue', fg='black')
                self.tempsgstp.place(x=567, y=self.y_index)
                labels_list.append(self.tempsgstp)
                self.y_index += 35
                self.counter += 1
                # total configure
                self.total_l.configure(text="Total: Rs." + str("%.2f"%sum(product_amount)))
                self.quantity_l.place_forget()
                self.quantity_e.place_forget()
                self.discount_l.place_forget()
                self.discount_e.place_forget()
                self.productname.configure(text="")
                #self.pprice.configure(text="") 
                self.add_to_cart_btn.destroy()
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,90)).place(x=670, y=90)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,125)).place(x=670, y=125)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,160)).place(x=670, y=160)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,195)).place(x=670, y=195)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,230)).place(x=670, y=230)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,265)).place(x=670, y=265)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,300)).place(x=670, y=300)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,335)).place(x=670, y=335)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,370)).place(x=670, y=370)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,405)).place(x=670, y=405)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,440)).place(x=670, y=440)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,475)).place(x=670, y=475)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,510)).place(x=670, y=510)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,545)).place(x=670, y=545)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,580)).place(x=670, y=580)
            Button(self.right, text="X", compound='center', font=("arial  12 bold"), bg='Blue', fg='white', width=4,height=1, command=lambda: account1(self.p,615)).place(x=670, y=615)
    def change_func(self, *args, **kwargs):
        self.amount_given = float(self.change_e.get())
        self.our_total = float(sum(product_amount))
        self.to_give = self.amount_given - self.our_total
        self.c_amount = Label(self.right, text="Change: Rs. " + str(self.to_give), font=('arial 18 bold'), fg='black', bg='red')
        self.c_amount.place(x=0 , y=600)
    def printers(self, *args, **kwargs):
        for die in range(0,len(f)):
            sck=(f[0])
            yi=(g[0])
            li=(h[0])
            il=(y[0])
        lio=(0,sck)
        kil="Mob No:"+str(il)
        con=sqlite3.Connection('vi.db')
        cur=con.cursor()
        cur.execute("insert into invoice values(?,?)",lio)
        con.commit()
        ree = cur.fetchall()
        #print(ree)
        query="SELECT rowid FROM invoice ORDER BY rowid DESC LIMIT 1"
        cur.execute(query)
        con.commit()
        reel = cur.fetchall()
        #print(reel)
        nk=""
        for row in reel:
            #print(row[0])
            nk=str(row[0])
        fontStyle = Font(name='Tahoma',size=10)
        fontamount = Font(name='Tahoma',size=12)
        ali = Alignment(horizontal='center',shrink_to_fit=True,indent=0)
        alia = Alignment(shrink_to_fit=True,indent=0)
        fontS = Font(name='Comic Sans MS',size=9)
        alie = Alignment(horizontal='center',shrink_to_fit=True,indent=0)
        now = datetime.datetime.now()
        hr=now.hour
        mn=now.minute
        if(str(mn)=='0' or str(mn)=='1'or str(mn)=='2'or str(mn)=='1'or str(mn)=='3'or str(mn)=='4'or str(mn)=='5'or str(mn)=='6'or str(mn)=='7'or str(mn)=='8'or str(mn)=='9'):
            mn='0'+str(mn)
        sc=now.second
        today = date.today()
        d1 = str(today.strftime("%d-%m-%Y"))
        directory = "D:/Store Management Software/Invoice/" + str(d1) + "/"
        if not os.path.exists(directory):
            os.makedirs(directory)
        wb=Workbook()
        a="00"+nk
        b="Date : "+str(d1)+"     "+str(hr)+":"+str(mn)
        
        file_name = str(directory) + str(a)+".xlsx"
        sheet=wb.active
        c='INVOICE NO. '+str(a)
        #sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_TABLOID
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        #pageSetup(wb, sheet = 2, orientation = "portrait", scale = 300, left= 0.5, right = 0.5)
        sheet.sheet_properties.fitToWidth = "FALSE"


        
        sheet.page_setup.fitToHeight = 1
        sheet.page_setup.fitToWidth = 0
        sheet.page_margins.left = 0
        sheet.page_margins.right = 0
        sheet.merge_cells('J2:N2')
        top_left_cell = sheet['J2']
        top_left_cell.value = str(c)
        top_left_cell.font = fontStyle
        top_left_cell.alignment = ali
        sheet.merge_cells('J3:N3')
        top_left_cell = sheet['J3']
        top_left_cell.value = str(b)
        top_left_cell.font = fontStyle
        sheet.merge_cells('J4:N4')
        top_left_cell = sheet['J4']
        top_left_cell.value = 'Liscense No:   20B/492/61/2019'
        top_left_cell.font = fontStyle
        sheet.merge_cells('J5:N5')
        top_left_cell = sheet['J5']
        top_left_cell.value = 'GUNA (M.P)'
        top_left_cell.font = fontStyle
        top_left_cell.alignment = ali
        sheet.merge_cells('A1:M1')
        top_left_cell = sheet['A1']
        top_left_cell.value = 'TAX INVOICE  -(CASH / CREDIT)'
        top_left_cell.font = fontStyle
        top_left_cell.alignment = ali
        sheet.merge_cells('D2:I2')
        top_left_cell = sheet['D2']
        top_left_cell.value = 'TO,'
        top_left_cell.font = fontStyle
        sheet.merge_cells('D3:I3')
        top_left_cell = sheet['D3']
        top_left_cell.value = str(sck)
        top_left_cell.font = fontStyle
        sheet.merge_cells('D4:I4')
        top_left_cell = sheet['D4']
        top_left_cell.value = str(yi)
        top_left_cell.font = fontStyle
        sheet.merge_cells('D5:I5')
        top_left_cell = sheet['D5']
        top_left_cell.value = str(kil)
        top_left_cell.font = fontStyle
        sheet.merge_cells('D6:I6')
        top_left_cell = sheet['D6']
        top_left_cell.value = str(li)
        top_left_cell.font = fontStyle
        sheet.merge_cells('A2:B2')
        top_left_cell = sheet['A2']
        top_left_cell.value = 'FROM,'
        top_left_cell.font = fontStyle
        sheet.merge_cells('A3:C3')
        top_left_cell = sheet['A3']
        top_left_cell.value = "     SHRINATH AGENCY -GUNA "
        top_left_cell.font = fontStyle
        sheet.merge_cells('A4:C4')
        top_left_cell = sheet['A4']
        top_left_cell.value = "     KOTESWAR MANDIR GALI "
        top_left_cell.font = fontStyle
        top_left_cell.font = fontStyle
        sheet.merge_cells('A5:C5')
        top_left_cell = sheet['A5']
        top_left_cell.value = '     MOB NO. 9993370326 '
        top_left_cell.font = fontStyle
        sheet.merge_cells('A6:C6')
        top_left_cell = sheet['A6']
        top_left_cell.value = '     GST NO.  23AQEPM5243K1ZU'
        top_left_cell.font = fontStyle
        sheet.column_dimensions['A'].width =5
        sheet.column_dimensions['B'].width =7
        sheet.column_dimensions['C'].width = 23
        sheet.column_dimensions['D'].width = 7
        sheet.column_dimensions['E'].width = 6
        sheet.column_dimensions['F'].width = 7
        sheet.column_dimensions['G'].width = 6
        sheet.column_dimensions['H'].width = 7
        sheet.column_dimensions['I'].width = 6
        sheet.column_dimensions['J'].width = 6
        sheet.column_dimensions['K'].width = 6
        sheet.column_dimensions['L'].width = 6
        sheet.column_dimensions['M'].width = 7
        sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=13)
        sheet.cell(row=7, column=1).value="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.cell(row=7, column=1).font = fontStyle
        sheet['A8']='S.NO'
        sheet['B8']='HSN CODE'
        sheet['C8']='Product'
        sheet['D8']='MRP'
        sheet['E8']='QTTY'
        sheet['F8']='RATE'
        sheet['G8']='Dscnt%'
        sheet['H8']='TAX AMT'
        sheet['I8']='CGST%'
        sheet['J8']='CGST'
        sheet['K8']='SGST%'
        sheet['L8']='SGST'
        sheet['M8']='AMOUNT'
        for  j in range(1,14):
            sheet.cell(row=8, column=j).font = fontStyle
            sheet.cell(row=8, column=j).alignment = ali
        sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=13)
        sheet.cell(row=9, column=1).value="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.cell(row=9, column=1).font = fontStyle    
        r = 1
        i = 0
        o=6
        l = 0
        i=0
        while l < len(products_list):
            sheet.cell(row=l+10, column=1).value = str(i+1)
            sheet.cell(row=l+10, column=2).value = str(product_code[i])
            sheet.cell(row=l+10, column=3).value = str(products_list[i])
            sheet.cell(row=l+10, column=4).value = str(product_mrp[i])
            sheet.cell(row=l+10, column=5).value = str(product_quantity[i])
            sheet.cell(row=l+10, column=6).value = str(product_rate[i])
            sheet.cell(row=l+10, column=7).value = str(product_dis[i])
            sheet.cell(row=l+10, column=8).value = str(product_price[i])
            sheet.cell(row=l+10, column=9).value = str(product_cgst[i])
            sheet.cell(row=l+10, column=10).value = str(product_cgstp[i])
            sheet.cell(row=l+10, column=11).value = str(product_sgst[i])
            sheet.cell(row=l+10, column=12).value = str(product_sgstp[i])
            sheet.cell(row=l+10, column=13).value = str(product_amount[i])
            sheet.cell(row=l+10, column=1).font = fontStyle
            sheet.cell(row=l+10, column=1).font = fontStyle
            sheet.cell(row=l+10, column=2).font = fontStyle
            sheet.cell(row=l+10, column=3).font = fontStyle
            sheet.cell(row=l+10, column=4).font = fontStyle
            sheet.cell(row=l+10, column=5).font = fontStyle
            sheet.cell(row=l+10, column=6).font = fontStyle
            sheet.cell(row=l+10, column=7).font = fontStyle
            sheet.cell(row=l+10, column=8).font = fontStyle
            sheet.cell(row=l+10, column=9).font = fontStyle
            sheet.cell(row=l+10, column=10).font = fontStyle
            sheet.cell(row=l+10, column=11).font = fontStyle
            sheet.cell(row=l+10, column=12).font = fontStyle
            sheet.cell(row=l+10, column=13).font = fontStyle
            sheet.cell(row=l+10, column=1).alignment = ali
            sheet.cell(row=l+10, column=2).alignment = ali
            sheet.cell(row=l+10, column=3).alignment = ali
            sheet.cell(row=l+10, column=4).alignment = ali
            sheet.cell(row=l+10, column=5).alignment = ali
            sheet.cell(row=l+10, column=6).alignment = ali
            sheet.cell(row=l+10, column=7).alignment = ali
            sheet.cell(row=l+10, column=8).alignment = ali
            sheet.cell(row=l+10, column=9).alignment = ali
            sheet.cell(row=l+10, column=10).alignment = ali
            sheet.cell(row=l+10, column=11).alignment = ali
            sheet.cell(row=l+10, column=12).alignment = ali
            sheet.cell(row=l+10, column=13).alignment = ali
            l=l+1
            i=i+1
        sheet.merge_cells(start_row=l+10, start_column=1, end_row=l+10, end_column=13)
        sheet.cell(row=l+10, column=1).value="------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"
        sheet.cell(row=l+10, column=1).font = fontStyle        
        sheet.cell(row=l+11, column=4).value = "TOTAL"
        sheet.cell(row=l+11, column=4).font = fontStyle
        amtw=0

        for row in sheet.iter_rows(min_row=10, min_col=13, max_row=l+10-1, max_col=13):
            for cell in row:
                amtw=amtw+float(cell.value)
                #print(amtw)
        sheet.cell(row=l+11, column=13).value = str("%.2f"%amtw)
        sheet.cell(row=l+11, column=13).font = fontStyle
        sheet.cell(row=l+11, column=13).alignment = ali
        qtyw=0
        for row in sheet.iter_rows(min_row=10, min_col=5, max_row=l+9, max_col=5):
            for cell in row:
                qtyw=qtyw+float(cell.value)
        sheet.cell(row=l+11, column=5).value = str("%.2f"%qtyw)
        sheet.cell(row=l+11, column=5).font = fontStyle
        sheet.cell(row=l+11, column=5).alignment = ali
        tyw=0
        for row in sheet.iter_rows(min_row=10, min_col=8, max_row=l+9, max_col=8):
            for cell in row:
                tyw=tyw+float(cell.value)
        sheet.cell(row=l+11, column=8).value = str("%.2f"%tyw)
        sheet.cell(row=l+11, column=8).font = fontStyle
        sheet.cell(row=l+11, column=8).alignment = ali
        yw=0
        for row in sheet.iter_rows(min_row=10, min_col=10, max_row=l+9, max_col=10):
            for cell in row:
                yw=yw+float(cell.value)
        sheet.cell(row=l+11, column=10).value = str("%.2f"%yw)
        sheet.cell(row=l+11, column=10).font = fontStyle
        sheet.cell(row=l+11, column=10).alignment = ali
        w=0
        for row in sheet.iter_rows(min_row=10, min_col=12, max_row=l+9, max_col=12):
            for cell in row:
                w=w+float(cell.value)
        sheet.cell(row=l+11, column=12).value = str("%.2f"%w)
        sheet.cell(row=l+11, column=12).alignment = ali
        sheet.cell(row=l+11, column=12).font = fontStyle       
        sheet.merge_cells(start_row=l+12, start_column=1, end_row=12+l, end_column=3)
        sheet.cell(row=l+12, column=1).value="TAX DETAILS"
        sheet.cell(row=l+12, column=1).alignment = ali
        sheet.cell(row=l+12, column=1).font = fontStyle
        sheet.merge_cells(start_row=l+13, start_column=1, end_row=13+l, end_column=3)
        sheet.cell(row=l+13, column=1).value="------------------------------------------------"
        sheet.cell(row=l+13, column=1).font = fontStyle        
        sheet.merge_cells(start_row=l+14, start_column=1, end_row=l+14, end_column=3)
        sheet.cell(row=l+14, column=1).value="Tax       Tax%        Taxable           TaxAmt"
        sheet.cell(row=l+14, column=1).font = fontStyle
        sheet.cell(row=l+14, column=1).alignment = ali
        sheet.cell(row=l+15, column=1).value = "CGST"
        sheet.cell(row=l+15, column=1).font = fontStyle
        sheet.cell(row=l+15, column=1).alignment=ali
        sheet.cell(row=l+16, column=1).value = "SGST"
        sheet.cell(row=l+16, column=1).font = fontStyle
        sheet.cell(row=l+16, column=1).alignment=ali
        sheet.cell(row=l+15, column=2).value = "0%"
        sheet.cell(row=l+15, column=1).font = fontStyle
        sheet.cell(row=l+15, column=2).alignment=ali
        sheet.cell(row=l+16, column=2).value = "0%"
        sheet.cell(row=l+16, column=1).font = fontStyle
        sheet.cell(row=l+16, column=2).alignment=ali
        txc=0
        txr=0
        di=0
        for di in range(0,len(product_cgst)):
            if product_cgst[di] == 0:
                txc=txc+product_price[di]
                txr =product_cgstp[di]+txr
            di=di+1
        sheet.cell(row=l+15, column=3).value = str("%.2f"%txc)+"            " +str("%.2f"%txr)
        sheet.cell(row=l+15, column=3).font = fontStyle
        sheet.cell(row=l+15, column=3).alignment=ali
        sheet.cell(row=l+16, column=3).value = str("%.2f"%txc)+"            "+str("%.2f"%txr)
        sheet.cell(row=l+16, column=3).font = fontStyle
        sheet.cell(row=l+16, column=3).alignment=ali        
        sheet.cell(row=l+17, column=1).value = "CGST"
        sheet.cell(row=l+17, column=1).font = fontStyle
        sheet.cell(row=l+17, column=1).alignment=ali
        sheet.cell(row=l+18, column=1).value = "SGST"
        sheet.cell(row=l+18, column=1).font = fontStyle
        sheet.cell(row=l+18, column=1).alignment=ali
        sheet.cell(row=l+17, column=2).value = "6%"
        sheet.cell(row=l+17, column=2).font = fontStyle
        sheet.cell(row=l+17, column=2).alignment=ali
        sheet.cell(row=l+18, column=2).value = "6%"
        sheet.cell(row=l+18, column=2).font = fontStyle
        sheet.cell(row=l+18, column=2).alignment=ali
        sheet.cell(row=l+19, column=1).value = "CGST"
        sheet.cell(row=l+19, column=1).font = fontStyle
        sheet.cell(row=l+19, column=1).alignment=ali
        sheet.cell(row=l+20, column=1).value = "SGST"
        sheet.cell(row=l+20, column=1).font = fontStyle
        sheet.cell(row=l+20, column=1).alignment=ali
        sheet.cell(row=l+19, column=2).value = "9%"
        sheet.cell(row=l+19, column=2).font = fontStyle
        sheet.cell(row=l+19, column=2).alignment=ali
        sheet.cell(row=l+20, column=2).value = "9%"
        sheet.cell(row=l+20, column=2).font = fontStyle
        sheet.cell(row=l+20, column=2).alignment=ali
        txc=0
        txr=0
        di=0

        for di in range(0,len(product_cgst)):
            if product_cgst[di] == 9:
                txc=txc+product_price[di]
                txr =product_cgstp[di]+txr
            di=di+1
        sheet.cell(row=l+19, column=3).value = str("%.2f"%txc)+"            " +str("%.2f"%txr)
        sheet.cell(row=l+19, column=3).font = fontStyle
        sheet.cell(row=l+19, column=3).alignment=ali
        sheet.cell(row=l+20, column=3).value = str("%.2f"%txc)+"            "+str("%.2f"%txr)
        sheet.cell(row=l+20, column=3).font = fontStyle
        sheet.cell(row=l+20, column=3).alignment=ali
        txr=0
        txc=0
        for di in range(0,len(product_cgst)):
            if product_cgst[di] == 6:
                txc=txc+product_price[di]
                txr =product_cgstp[di]+txr
            di=di+1
        sheet.cell(row=l+17, column=3).value = str("%.2f"%txc) +"            "+str("%.2f"%txr)
        sheet.cell(row=l+17, column=3).font = fontStyle
        sheet.cell(row=l+17, column=3).alignment=ali
        sheet.cell(row=l+18, column=3).value = str("%.2f"%txc) +"            "+str("%.2f"%txr)
        sheet.cell(row=l+18, column=3).font = fontStyle
        sheet.cell(row=l+18, column=3).alignment=ali
        sheet.merge_cells(start_row=l+14, start_column=10, end_row=l+14, end_column=12)
        sheet.cell(row=l+14, column=10).value="Gross Amt"
        tyiw=float(sum(akm))
        sheet.cell(row=l+14, column=10).font = fontStyle        
        sheet.merge_cells(start_row=l+14, start_column=13, end_row=l+14, end_column=13)
        sheet.cell(row=l+14, column=13).value=str("%.2f"%tyiw)
        sheet.cell(row=l+14, column=13).alignment = ali
        sheet.cell(row=l+14, column=13).font = fontStyle        
        k=[]
        for i in range(0,len(product_dis)):
            k.append((float(product_dis[i])*product_rate[i]*product_quantity[i])/100.0)
        discountvalue=float("%.2f"%sum(k))         
        sheet.merge_cells(start_row=l+15, start_column=10, end_row=l+15, end_column=12)
        sheet.cell(row=l+15, column=10).value="Scheme Disc Amt(-)"
        sheet.cell(row=l+15, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+15, start_column=13, end_row=l+15, end_column=13)
        sheet.cell(row=l+15, column=13).value=discountvalue
        sheet.cell(row=l+15, column=13).alignment = ali
        sheet.cell(row=l+15, column=13).font = fontStyle
        taxamts=(sum(product_cgstp)+sum(product_sgstp))
        sheet.merge_cells(start_row=l+16, start_column=10, end_row=l+16, end_column=12)
        sheet.cell(row=l+16, column=10).value="Tax Amt(+)"
        sheet.cell(row=l+16, column=10).font = fontStyle

        sheet.merge_cells(start_row=l+16, start_column=13, end_row=l+16, end_column=13)
        sheet.cell(row=l+16, column=13).value=taxamts
        sheet.cell(row=l+16, column=13).alignment = ali
        sheet.cell(row=l+16, column=13).font = fontStyle
        totals=tyiw + taxamts - discountvalue
        jss=round(totals)
        js=str(jss)+".00"
        lml=(jss-totals)
        lnl=str(round(lml,2))    
        sheet.merge_cells(start_row=l+17, start_column=10, end_row=l+17, end_column=12)
        sheet.cell(row=l+17, column=10).value="Round Off"
        sheet.cell(row=l+17, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+17, start_column=13, end_row=l+17, end_column=13)
        sheet.cell(row=l+17, column=13).value=lnl
        sheet.cell(row=l+17, column=13).alignment = ali
        sheet.cell(row=l+17, column=13).font = fontStyle        
        sheet.merge_cells(start_row=l+18, start_column=10, end_row=l+18, end_column=12)
        sheet.cell(row=l+18, column=10).value="Net Payable"
        sheet.cell(row=l+18, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+18, start_column=13, end_row=l+18, end_column=14)
        sheet.cell(row=l+18, column=13).value=js
        sheet.cell(row=l+18, column=13).alignment = alia
        sheet.cell(row=l+18, column=13).font = fontamount
        sheet.merge_cells(start_row=l+19, start_column=10, end_row=l+19, end_column=13)
        sheet.cell(row=l+19, column=10).value="From Shrinath Agency Guna (M.P)"
        sheet.cell(row=l+19, column=10).alignment = ali
        sheet.cell(row=l+19, column=10).font = fontStyle
        sheet.merge_cells(start_row=l+21, start_column=1, end_row=l+21, end_column=13)
        sheet.cell(row=l+21, column=1).value="DECLARATION:We here by certify that our rc under the GST ACT is in force                        Bill created by SCYIL 9009323236"
        sheet.cell(row=l+21, column=1).alignment = ali
        sheet.cell(row=l+21, column=1).font = fontStyle
        sheet.merge_cells(start_row=l+22, start_column=1, end_row=l+22, end_column=10)
        sheet.cell(row=l+22, column=1).value="on the date on which the sale of goods specified in this tax invoice"
        sheet.cell(row=l+22, column=1).alignment = ali
        sheet.cell(row=l+22, column=1).font = fontStyle   
        sheet.merge_cells(start_row=l+25, start_column=1, end_row=l+25, end_column=13)
        sheet.cell(row=l+25, column=1).value="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------"

        wb.save(file_name)
        os.startfile(file_name,"print")
root = Tk()
b = Application(root)
root.geometry("1366x768+0+0")
root.mainloop()
